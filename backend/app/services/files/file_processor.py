"""
File processing module for extracting text from uploaded documents.
Spreadsheets (xlsx/csv/tsv) are summarised locally; images return metadata
only; everything else (documents, presentations, web files) goes through
LlamaParse.
"""
import io
import logging
from dataclasses import dataclass
from typing import Any, Awaitable, Callable, Dict, List, Optional, Tuple

from fastapi import HTTPException, UploadFile
from llama_cloud import AsyncLlamaCloud
from PIL import Image

from ...config.settings import settings
from ...config.file_types import (
    ALL_EXTENSIONS,
    MIME_TO_EXTENSION,
    IMAGE_EXTENSIONS,
    SPREADSHEET_EXTENSIONS,
)
from ...logging import log_event
from ..pii_redactor import UserRedactionPatterns, redact_filename, redact_text
from .staged_upload_cancellation_service import StagedUploadCancelledError

logger = logging.getLogger(__name__)

# Spreadsheet formats we can summarise locally (no LlamaParse needed)
_LOCAL_SPREADSHEET_TYPES = {"xlsx", "xlsm", "xlsb", "csv", "tsv"}
_BYTE_REDACTION_SUPPORTED_SPREADSHEET_TYPES = {"xlsx", "xlsm", "csv", "tsv"}


@dataclass(frozen=True)
class SpreadsheetByteRedactionResult:
    file_content: bytes
    redaction_performed: bool
    redaction_hits: List[str]


class FileProcessor:
    """Handles text extraction from 80+ file formats via LlamaParse."""

    @classmethod
    def is_spreadsheet_type(cls, file_type: str) -> bool:
        return str(file_type or "").strip().lower() in SPREADSHEET_EXTENSIONS

    @classmethod
    def supports_spreadsheet_byte_redaction(cls, file_type: str) -> bool:
        return str(file_type or "").strip().lower() in _BYTE_REDACTION_SUPPORTED_SPREADSHEET_TYPES

    @staticmethod
    async def _raise_if_cancelled(
        cancel_check: Optional[Callable[[], Awaitable[bool]]],
    ) -> None:
        if cancel_check is not None and bool(await cancel_check()):
            raise StagedUploadCancelledError("Upload cancelled by user.")
    
    @classmethod
    def detect_file_type(cls, file: UploadFile) -> str:
        """
        Detect file type from MIME type and extension.
        Returns extension string (e.g., 'pdf', 'docx') or 'unsupported'.
        """
        # Check MIME type first
        if file.content_type and file.content_type in MIME_TO_EXTENSION:
            return MIME_TO_EXTENSION[file.content_type]

        # Fallback to extension
        if file.filename:
            parts = file.filename.lower().split(".")
            if len(parts) > 1:
                extension = parts[-1]
                if extension in ALL_EXTENSIONS:
                    return extension

        return "unsupported"
    
    @classmethod
    async def validate_file(cls, file: UploadFile) -> str:
        """Validate file type and return detected extension string."""
        file_type = cls.detect_file_type(file)
        if file_type == "unsupported":
            supported_types = ", ".join(settings.allowed_file_types)
            raise HTTPException(
                status_code=400,
                detail=f"Unsupported file type. Supported formats: {supported_types}"
            )

        return file_type

    @classmethod
    async def extract_text_with_llamaparse(
        cls,
        file_content: bytes,
        file_type: str,
        filename: str,
        cancel_check: Optional[Callable[[], Awaitable[bool]]] = None,
    ) -> str:
        """
        Extract text from documents using LlamaParse.
        Supports 70+ formats including documents, spreadsheets, presentations, images, and web files.
        """
        if not settings.llama_cloud_api_key:
            log_event(
                logger,
                "ERROR",
                "file.parse.llamaparse_key_missing",
                "error",
                file_type=file_type,
            )
            return f"LlamaParse API key not configured. Cannot extract text from {file_type.upper()} file."

        try:
            await cls._raise_if_cancelled(cancel_check)
            base_url = str(getattr(settings, "llama_cloud_base_url", "") or "").strip() or None
            log_event(
                logger,
                "INFO",
                "file.parse.llamaparse.start",
                "timing",
                file_type=file_type,
                filename=filename,
            )

            upload_name = filename or f"document.{file_type}"
            async with AsyncLlamaCloud(
                api_key=settings.llama_cloud_api_key,
                base_url=base_url,
            ) as client:
                parsed = await client.parsing.parse(
                    tier="cost_effective",
                    version="latest",
                    upload_file=(upload_name, file_content),
                    expand=["markdown"],
                    timeout=300.0,
                )
            await cls._raise_if_cancelled(cancel_check)

            text = str(getattr(parsed, "markdown_full", "") or "").strip()
            if not text:
                markdown = getattr(parsed, "markdown", None)
                pages = list(getattr(markdown, "pages", []) or [])
                rendered_pages: List[str] = []
                for page in pages:
                    page_markdown = str(getattr(page, "markdown", "") or "").strip()
                    if page_markdown:
                        rendered_pages.append(page_markdown)
                text = "\n\n---\n\n".join(rendered_pages).strip()

            if not text:
                log_event(
                    logger,
                    "WARNING",
                    "file.parse.llamaparse.empty",
                    "retry",
                    file_type=file_type,
                    filename=filename,
                )
                return f"No readable text content found in {file_type.upper()} file."

            log_event(
                logger,
                "INFO",
                "file.parse.llamaparse.success",
                "timing",
                file_type=file_type,
                filename=filename,
                char_count=len(text),
            )
            return text

        except StagedUploadCancelledError:
            raise
        except Exception as exc:
            log_event(
                logger,
                "ERROR",
                "file.parse.llamaparse.failed",
                "error",
                file_type=file_type,
                filename=filename,
                exc_info=exc,
            )
            return f"Failed to extract text from {file_type.upper()} file: {str(exc)}"

    @classmethod
    async def extract_text_from_txt(cls, file_content: bytes) -> str:
        """Extract text from TXT file."""
        try:
            # Try UTF-8 first, fallback to latin-1
            try:
                text = file_content.decode('utf-8')
            except UnicodeDecodeError:
                text = file_content.decode('latin-1')

            if not text.strip():
                return "No readable text content found in TXT file."

            return text.strip()

        except Exception as exc:
            log_event(
                logger,
                "ERROR",
                "file.parse.txt.failed",
                "error",
                exc_info=exc,
            )
            return f"Failed to extract text from TXT: {str(exc)}"

    @classmethod
    def _normalize_image(cls, image_bytes: bytes, source_format: str = "unknown") -> Optional[Tuple[bytes, str]]:
        """
        Normalize and validate image bytes, converting to standard formats if needed.
        Resizes images to max 768px per dimension to reduce context usage.
        Returns tuple of (normalized_bytes, file_extension) or None if invalid.
        """
        try:
            img = Image.open(io.BytesIO(image_bytes))

            # Skip very small images (likely icons or noise)
            if img.width < 32 or img.height < 32:
                return None

            # Resize if either dimension exceeds 768 pixels (context optimization)
            max_dimension = 768
            if img.width > max_dimension or img.height > max_dimension:
                # Calculate new dimensions maintaining aspect ratio
                if img.width > img.height:
                    new_width = max_dimension
                    new_height = int(img.height * (max_dimension / img.width))
                else:
                    new_height = max_dimension
                    new_width = int(img.width * (max_dimension / img.height))

                # Use LANCZOS for high-quality downsampling
                img = img.resize((new_width, new_height), Image.Resampling.LANCZOS)
                log_event(
                    logger,
                    "INFO",
                    "file.image.normalized_resize",
                    "timing",
                    width=new_width,
                    height=new_height,
                )

            # Convert to RGB if necessary (handles RGBA, P, L modes)
            if img.mode not in ("RGB", "L"):
                if img.mode == "RGBA":
                    # Create white background
                    background = Image.new("RGB", img.size, (255, 255, 255))
                    background.paste(img, mask=img.split()[3] if len(img.split()) == 4 else None)
                    img = background
                else:
                    img = img.convert("RGB")

            # Determine output format
            output_format = "JPEG"
            extension = "jpg"
            if img.mode == "L":  # Grayscale
                output_format = "JPEG"
                extension = "jpg"
            elif source_format.lower() in ["png", "image/png"]:
                output_format = "PNG"
                extension = "png"

            # Re-encode to standardized format
            output_buffer = io.BytesIO()
            img.save(output_buffer, format=output_format, quality=85 if output_format == "JPEG" else None)
            normalized_bytes = output_buffer.getvalue()

            return (normalized_bytes, extension)

        except Exception as exc:
            log_event(
                logger,
                "DEBUG",
                "file.image.normalize_failed",
                "retry",
                source_format=source_format,
                error=str(exc),
            )
            return None

    @classmethod
    async def extract_text_from_image(cls, file_content: bytes, file_type: str) -> str:
        """Handle image files - return metadata instead of text."""
        try:
            file_size_kb = len(file_content) / 1024
            return f"Image file ({file_type.upper()}) - {file_size_kb:.1f}KB. Visual content cannot be extracted as text."

        except Exception as exc:
            log_event(
                logger,
                "ERROR",
                "file.image.process_failed",
                "error",
                file_type=file_type,
                exc_info=exc,
            )
            return f"Failed to process {file_type.upper()} image: {str(exc)}"
    
    @classmethod
    async def extract_text_from_spreadsheet(
        cls,
        file_content: bytes,
        file_type: str,
        filename: str,
        cancel_check: Optional[Callable[[], Awaitable[bool]]] = None,
    ) -> str:
        """Generate a compact metadata summary for spreadsheet files.

        Uses openpyxl (xlsx/xlsm/xlsb) or the csv module (csv/tsv) so we
        never need to call LlamaParse for data files.  The summary is small
        enough to fit in context while giving the model enough info to write
        targeted pandas code in the sandbox.
        """
        MAX_SAMPLE_ROWS = 5
        MAX_COLUMNS_SHOWN = 30

        try:
            if file_type in ("xlsx", "xlsm", "xlsb"):
                return await cls._summarise_excel(file_content, filename, MAX_SAMPLE_ROWS, MAX_COLUMNS_SHOWN)
            elif file_type in ("csv", "tsv"):
                return cls._summarise_csv(file_content, file_type, filename, MAX_SAMPLE_ROWS, MAX_COLUMNS_SHOWN)
            else:
                # Shouldn't reach here, but fall back to LlamaParse
                return await cls.extract_text_with_llamaparse(
                    file_content,
                    file_type,
                    filename,
                    cancel_check=cancel_check,
                )
        except StagedUploadCancelledError:
            raise
        except Exception as exc:
            log_event(
                logger,
                "WARNING",
                "file.parse.spreadsheet_summary_failed",
                "retry",
                filename=filename,
                file_type=file_type,
                error=str(exc),
            )
            return await cls.extract_text_with_llamaparse(
                file_content,
                file_type,
                filename,
                cancel_check=cancel_check,
            )

    @classmethod
    async def _summarise_excel(cls, file_content: bytes, filename: str, max_rows: int, max_cols: int) -> str:
        """Summarise an Excel workbook using openpyxl (read-only mode)."""
        from anyio import to_thread

        def _read():
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
            parts: list[str] = [f"# {filename}\n"]

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(max_row=max_rows + 1, values_only=True))

                if not rows:
                    parts.append(f"## Sheet: {sheet_name}\n(empty sheet)\n")
                    continue

                # First row = headers
                headers = [str(c) if c is not None else "" for c in rows[0]]
                total_cols = len(headers)
                truncated_cols = total_cols > max_cols
                headers = headers[:max_cols]

                # Count total rows (read-only mode has ws.max_row)
                total_rows = (ws.max_row or 1) - 1  # subtract header

                parts.append(f"## Sheet: {sheet_name}")
                parts.append(f"Rows: ~{total_rows:,} | Columns: {total_cols}")
                col_display = ", ".join(headers)
                if truncated_cols:
                    col_display += f", ... (+{total_cols - max_cols} more)"
                parts.append(f"Columns: {col_display}\n")

                # Sample rows as a markdown table
                if len(rows) > 1:
                    sample = rows[1:max_rows + 1]
                    # Header row
                    hdr_line = "| " + " | ".join(h or "(blank)" for h in headers) + " |"
                    sep_line = "| " + " | ".join("---" for _ in headers) + " |"
                    parts.append(hdr_line)
                    parts.append(sep_line)
                    for row in sample:
                        cells = [str(c) if c is not None else "" for c in row[:max_cols]]
                        # Truncate individual cell values to keep summary lean
                        cells = [c[:60] + "…" if len(c) > 60 else c for c in cells]
                        parts.append("| " + " | ".join(cells) + " |")
                    if total_rows > max_rows:
                        parts.append(f"\n... ({total_rows - max_rows:,} more rows)")

                parts.append("")  # blank line between sheets

            wb.close()
            return "\n".join(parts)

        return await to_thread.run_sync(_read)

    @classmethod
    def _summarise_csv(cls, file_content: bytes, file_type: str, filename: str, max_rows: int, max_cols: int) -> str:
        """Summarise a CSV/TSV file using the csv module."""
        import csv as csv_module

        try:
            text = file_content.decode("utf-8")
        except UnicodeDecodeError:
            text = file_content.decode("latin-1")

        delimiter = "\t" if file_type == "tsv" else ","
        reader = csv_module.reader(text.splitlines(), delimiter=delimiter)

        rows = []
        total_rows = 0
        for i, row in enumerate(reader):
            if i <= max_rows:
                rows.append(row)
            total_rows = i

        if not rows:
            return f"# {filename}\n(empty file)"

        headers = rows[0]
        total_cols = len(headers)
        truncated_cols = total_cols > max_cols
        headers_display = headers[:max_cols]
        data_rows = total_rows  # subtract header row

        parts: list[str] = [f"# {filename}"]
        parts.append(f"Format: {file_type.upper()} | Rows: ~{data_rows:,} | Columns: {total_cols}")
        col_display = ", ".join(headers_display)
        if truncated_cols:
            col_display += f", ... (+{total_cols - max_cols} more)"
        parts.append(f"Columns: {col_display}\n")

        # Sample rows as a markdown table
        if len(rows) > 1:
            hdr_line = "| " + " | ".join(h or "(blank)" for h in headers_display) + " |"
            sep_line = "| " + " | ".join("---" for _ in headers_display) + " |"
            parts.append(hdr_line)
            parts.append(sep_line)
            for row in rows[1:max_rows + 1]:
                cells = [str(c) if c is not None else "" for c in row[:max_cols]]
                cells = [c[:60] + "…" if len(c) > 60 else c for c in cells]
                parts.append("| " + " | ".join(cells) + " |")
            if data_rows > max_rows:
                parts.append(f"\n... ({data_rows - max_rows:,} more rows)")

        return "\n".join(parts)

    @classmethod
    async def redact_spreadsheet_file_content(
        cls,
        file_content: bytes,
        file_type: str,
        *,
        user_redaction_list: Optional[List[str]] = None,
    ) -> SpreadsheetByteRedactionResult:
        """Apply redaction directly to spreadsheet bytes.

        This guarantees that downstream tools (`execute_code`) receive redacted
        spreadsheet bytes when upload redaction is enabled.
        """
        normalized_type = str(file_type or "").strip().lower()
        entries = [item for item in (user_redaction_list or []) if isinstance(item, str) and item.strip()]
        if not entries:
            return SpreadsheetByteRedactionResult(
                file_content=file_content,
                redaction_performed=False,
                redaction_hits=[],
            )

        if normalized_type in {"csv", "tsv"}:
            return await cls._redact_delimited_spreadsheet_bytes(
                file_content=file_content,
                file_type=normalized_type,
                user_redaction_list=entries,
            )

        if normalized_type in {"xlsx", "xlsm"}:
            return await cls._redact_excel_workbook_bytes(
                file_content=file_content,
                file_type=normalized_type,
                user_redaction_list=entries,
            )

        raise ValueError(
            "Redaction for this spreadsheet format is not supported yet. "
            "Use XLSX, XLSM, CSV, or TSV for redacted uploads."
        )

    @classmethod
    async def _redact_delimited_spreadsheet_bytes(
        cls,
        *,
        file_content: bytes,
        file_type: str,
        user_redaction_list: List[str],
    ) -> SpreadsheetByteRedactionResult:
        encoding_used = "utf-8"
        try:
            text = file_content.decode("utf-8")
        except UnicodeDecodeError:
            text = file_content.decode("latin-1")
            encoding_used = "latin-1"

        redaction_result = await redact_text(
            text,
            user_redaction_list=user_redaction_list,
        )
        if not redaction_result.redaction_performed:
            return SpreadsheetByteRedactionResult(
                file_content=file_content,
                redaction_performed=False,
                redaction_hits=[],
            )

        try:
            redacted_bytes = redaction_result.text.encode(encoding_used)
        except UnicodeEncodeError:
            redacted_bytes = redaction_result.text.encode("utf-8")

        return SpreadsheetByteRedactionResult(
            file_content=redacted_bytes,
            redaction_performed=True,
            redaction_hits=sorted(redaction_result.redaction_hits),
        )

    @classmethod
    async def _redact_excel_workbook_bytes(
        cls,
        *,
        file_content: bytes,
        file_type: str,
        user_redaction_list: List[str],
    ) -> SpreadsheetByteRedactionResult:
        from anyio import to_thread

        keep_vba = file_type == "xlsm"

        def _redact_workbook() -> SpreadsheetByteRedactionResult:
            import openpyxl
            from openpyxl.comments import Comment

            workbook = openpyxl.load_workbook(
                io.BytesIO(file_content),
                read_only=False,
                data_only=False,
                keep_vba=keep_vba,
            )
            try:
                hits: set[str] = set()
                changed = False

                for sheet in workbook.worksheets:
                    for row in sheet.iter_rows():
                        for cell in row:
                            value = cell.value
                            if isinstance(value, str) and value:
                                redacted_value, value_hits = UserRedactionPatterns.redact(
                                    value,
                                    user_redaction_list,
                                )
                                if value_hits:
                                    hits.update(value_hits)
                                    if redacted_value != value:
                                        cell.value = redacted_value
                                        changed = True

                            comment = getattr(cell, "comment", None)
                            if comment and isinstance(comment.text, str) and comment.text:
                                redacted_comment, comment_hits = UserRedactionPatterns.redact(
                                    comment.text,
                                    user_redaction_list,
                                )
                                if comment_hits:
                                    hits.update(comment_hits)
                                    if redacted_comment != comment.text:
                                        cell.comment = Comment(
                                            text=redacted_comment,
                                            author=comment.author or "assistant",
                                        )
                                        changed = True

                if not hits or not changed:
                    return SpreadsheetByteRedactionResult(
                        file_content=file_content,
                        redaction_performed=False,
                        redaction_hits=[],
                    )

                output = io.BytesIO()
                workbook.save(output)
                return SpreadsheetByteRedactionResult(
                    file_content=output.getvalue(),
                    redaction_performed=True,
                    redaction_hits=sorted(hits),
                )
            finally:
                workbook.close()

        return await to_thread.run_sync(_redact_workbook)

    @classmethod
    async def extract_text(
        cls,
        file_content: bytes,
        file_type: str,
        filename: str = "document",
        cancel_check: Optional[Callable[[], Awaitable[bool]]] = None,
    ) -> str:
        """
        Extract text content based on file type.
        Routes files to appropriate extraction method:
        - TXT: Direct text extraction
        - Images (PNG, JPG, etc.): Metadata only
        - Common spreadsheets (CSV/XLS/XLSX/TSV): Local metadata summary
        - Everything else: LlamaParse (documents, presentations, web, uncommon formats)
        """
        # Handle plain text directly (no need for LlamaParse)
        if file_type == "txt":
            return await cls.extract_text_from_txt(file_content)

        # Handle direct image uploads (metadata only, not processed by LlamaParse)
        if file_type in IMAGE_EXTENSIONS:
            return await cls.extract_text_from_image(file_content, file_type)

        # Spreadsheets we can summarise locally (skip LlamaParse, save API credits)
        if file_type in _LOCAL_SPREADSHEET_TYPES:
            return await cls.extract_text_from_spreadsheet(
                file_content,
                file_type,
                filename,
                cancel_check=cancel_check,
            )

        # Everything else goes through LlamaParse
        # This includes: documents, presentations, web files, and rare spreadsheet formats
        return await cls.extract_text_with_llamaparse(
            file_content,
            file_type,
            filename,
            cancel_check=cancel_check,
        )
    
    @classmethod
    async def process_file(
        cls,
        file: UploadFile,
        *,
        redact: bool = False,
        user_redaction_list: Optional[List[str]] = None,
        cancel_check: Optional[Callable[[], Awaitable[bool]]] = None,
    ) -> dict:
        """
        Process a single file and extract its text content.
        Returns dict with file info and extracted text.
        
        Args:
            file: The uploaded file to process.
            redact: Whether to apply PII redaction.
            user_redaction_list: Optional list of names/terms from user's custom redaction list.
        """
        await cls._raise_if_cancelled(cancel_check)
        file_type = await cls.validate_file(file)
        file_content = await file.read()

        content_size = len(file_content)

        # File size limit
        max_size = settings.max_file_size
        if max_size and content_size > max_size:
            raise HTTPException(
                status_code=413,
                detail=f"File size exceeds maximum allowed size of {max_size / 1024 / 1024:.0f}MB",
            )

        # Reset file pointer for potential re-use
        await file.seek(0)
        await cls._raise_if_cancelled(cancel_check)

        # Normalize standalone image files to ensure they meet Anthropic's dimension limits
        if file_type in IMAGE_EXTENSIONS:
            from anyio import to_thread
            normalized = await to_thread.run_sync(cls._normalize_image, file_content, file_type)
            if normalized:
                file_content, file_ext = normalized
                # Update file type if normalization changed the format
                file_type = file_ext
                content_size = len(file_content)
                log_event(
                    logger,
                    "INFO",
                    "file.image.normalized",
                    "timing",
                    filename=file.filename,
                    file_type=file_type,
                )

        spreadsheet_redaction_performed = False
        spreadsheet_redaction_hits: List[str] = []
        if redact and cls.is_spreadsheet_type(file_type):
            if not cls.supports_spreadsheet_byte_redaction(file_type):
                raise ValueError(
                    "Redaction for this spreadsheet format is not supported yet. "
                    "Use XLSX, XLSM, CSV, or TSV for redacted uploads."
                )
            spreadsheet_redaction_result = await cls.redact_spreadsheet_file_content(
                file_content,
                file_type,
                user_redaction_list=user_redaction_list,
            )
            file_content = spreadsheet_redaction_result.file_content
            content_size = len(file_content)
            spreadsheet_redaction_performed = spreadsheet_redaction_result.redaction_performed
            spreadsheet_redaction_hits = spreadsheet_redaction_result.redaction_hits
            if spreadsheet_redaction_performed:
                log_event(
                    logger,
                    "INFO",
                    "redaction.spreadsheet_bytes_applied",
                    "timing",
                    filename=file.filename,
                    file_type=file_type,
                    redaction_hits=spreadsheet_redaction_hits,
                )

        original_filename = file.filename or "unknown"
        filename_redaction_performed = False
        filename_redaction_hits: List[str] = []
        if redact:
            filename_result = redact_filename(
                original_filename,
                user_redaction_list=user_redaction_list,
            )
            original_filename = filename_result.text
            filename_redaction_performed = filename_result.redaction_performed
            filename_redaction_hits = filename_result.redaction_hits

        # Extract text content using LlamaParse for documents
        extracted_text = await cls.extract_text(
            file_content,
            file_type,
            original_filename or "document",
            cancel_check=cancel_check,
        )
        await cls._raise_if_cancelled(cancel_check)

        content_redaction_performed = False
        content_redaction_hits: List[str] = []

        # Apply redaction to text-based files (skip images which don't have extractable text)
        await cls._raise_if_cancelled(cancel_check)
        if redact and file_type not in IMAGE_EXTENSIONS:
            result = await redact_text(
                extracted_text,
                user_redaction_list=user_redaction_list,
            )
            extracted_text = result.text
            content_redaction_performed = result.redaction_performed
            content_redaction_hits = result.redaction_hits
            if content_redaction_performed:
                log_event(
                    logger,
                    "DEBUG",
                    "redaction.applied",
                    "timing",
                    file_name=file.filename,
                    file_type=file_type,
                    redaction_hits=content_redaction_hits,
                )
        await cls._raise_if_cancelled(cancel_check)

        redaction_hits = sorted(
            set(content_redaction_hits + filename_redaction_hits + spreadsheet_redaction_hits)
        )
        redaction_performed = (
            content_redaction_performed
            or filename_redaction_performed
            or spreadsheet_redaction_performed
        )

        return {
            "file_type": file_type,
            "original_filename": original_filename,
            "file_content": file_content,  # Normalized/redacted bytes for blob storage
            "extracted_text": extracted_text,
            "file_size": content_size,
            "char_count": len(extracted_text) if extracted_text else 0,
            "word_count": len(extracted_text.split()) if extracted_text else 0,
            "redaction_requested": redact,
            "redaction_performed": redaction_performed,
            "redaction_hits": redaction_hits,
            "filename_redaction_performed": filename_redaction_performed,
            "filename_redaction_hits": filename_redaction_hits,
            "spreadsheet_redaction_performed": spreadsheet_redaction_performed,
            "spreadsheet_redaction_hits": spreadsheet_redaction_hits,
            # Compatibility keys expected by API response models/clients.
            "user_redaction_performed": redaction_performed,
            "user_redaction_hits": redaction_hits,
            "redacted_categories": ["user_redaction"] if redaction_performed else [],
        }
